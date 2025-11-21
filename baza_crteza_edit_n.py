import os
import re
import tkinter as tk
from tkinter import ttk, messagebox
import tkinter.font as tkFont
import webbrowser
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
from datetime import datetime

def load_config():
    config_file = "config.txt"
    config = {}
    
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    if '=' in line:
                        key, value = line.split('=', 1)
                        config[key.strip()] = value.strip()
        except Exception as e:
            print(f"Error reading config.txt: {e}")
    
    return config

config = load_config()
BASE_FOLDER = config.get("BASE_FOLDER", ".")
EXCEL_FILENAME = config.get("EXCEL_FILENAME", "BAZACRTEZA.xlsx")

EXCEL_FILE = os.path.join(BASE_FOLDER, EXCEL_FILENAME)
DRAWINGS_FOLDER = os.path.join(BASE_FOLDER, "crtezi")
BACKUP_FOLDER = config.get("BACKUP_FOLDER", os.path.join(".", "backup"))

FIELDS = [
    "IDENTBROJ", "CRTEZBROJ", "NAZIVDELA", "TEHNPODACI",
    "KATALBROJ", "FORMAT", "ARHIVA", "KOMENTAR",
    "OBJEKAT1", "OBJEKAT2", "OBJEKAT3", "OBJEKAT4",
    "KATALOG", "MAGSIFRA"
]

# --- Load Excel with openpyxl ---
def load_excel_data(filename):
    """Load Excel data using openpyxl instead of pandas"""
    try:
        wb = load_workbook(filename, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Load data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                # Handle different data types
                if value is None or value == "":
                    record[header] = None
                else:
                    record[header] = value
            data.append(record)
        
        wb.close()
        return data, None
    except FileNotFoundError:
        return [], f"{filename} nije pronađen!"
    except Exception as e:
        return [], f"Greška pri učitavanju {filename}: {str(e)}"

def save_excel_data(filename, data, fields):
    """Save data to Excel using openpyxl"""
    try:
        wb = Workbook()
        ws = wb.active
        
        # Write headers
        for col, field in enumerate(fields, start=1):
            ws.cell(row=1, column=col, value=field)
        
        # Write data rows
        for row_idx, record in enumerate(data, start=2):
            for col_idx, field in enumerate(fields, start=1):
                value = record.get(field)
                # Handle None values
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filename)
        return True, None
    except PermissionError:
        return False, (
            f"Ne mogu da sačuvam '{filename}'!\n\n"
            f"Fajl je verovatno otvoren u Excel programu.\n"
            f"Molim vas zatvorite fajl u Excel-u i pokušajte ponovo."
        )
    except Exception as e:
        return False, f"Greška: {str(e)}"

data, error = load_excel_data(EXCEL_FILE)

current_index = None
search_results = []
search_index = 0

# --- Tkinter setup ---
root = tk.Tk()
root.title("Baza Crteza Pomoćne mehanizacije - Editor - Verzija 1.0 - M.Petković")
root.geometry("1400x800")

# Show error if data loading failed
if error:
    root.withdraw()
    messagebox.showerror("Greška", error)
    if not data:
        root.destroy()
        exit()
    root.deiconify()

record_number_var = tk.StringVar()

# --- Fonts ---
default_font = tkFont.nametofont("TkDefaultFont")
default_font.configure(family="Arial", size=12)
root.option_add("*Font", default_font)

# --- Colors ---
app_bg = "#e8eff0"
frame_bg = "#f6f9f9"

root.configure(bg=app_bg)

# --- Styles ---
style = ttk.Style()
style.theme_use("clam")

style.configure("TFrame", background=frame_bg, bordercolor="#dae4e6")
style.configure("TLabel", background=frame_bg)
style.configure("TButton", font=("Arial", 11, "bold"), padding=(10, 5))
style.configure("TEntry", fieldbackground="#ffffff", foreground="#000000", 
                bordercolor="#dae4e6", padding=(5, 2))

title_label = ttk.Label(root, text="Baza Crteža Pomoćne Mehanizacije - Editor", 
                        font=("Arial", 18, "bold"), background=app_bg, 
                        foreground="#333333", anchor="center")
title_label.pack(side="top", fill="x", pady=10)

# --- Search frame ---
search_frame = ttk.Frame(root, relief="flat", borderwidth=1)
search_frame.pack(fill=tk.X, padx=10, pady=5)

search_field_var = tk.StringVar(value="CRTEZBROJ")
search_value_var = tk.StringVar()

inner_frame = ttk.Frame(search_frame)
inner_frame.pack()

ttk.Label(inner_frame, text="Traži u:").pack(side=tk.LEFT, padx=0)
search_field_menu = ttk.Combobox(inner_frame, textvariable=search_field_var, 
                                  values=FIELDS, width=15, state="readonly")
search_field_menu.pack(side=tk.LEFT, padx=5)
ttk.Label(inner_frame, text="Vrednost:").pack(side=tk.LEFT, padx=(15, 0))
search_value_entry = ttk.Entry(inner_frame, textvariable=search_value_var, 
                               width=30, justify='center')
search_value_entry.pack(side=tk.LEFT, padx=0, pady=5, ipadx=5, ipady=8)

def normalize(s):
    """Normalize string for comparison"""
    if not s:
        return ""
    return re.sub(r"\W+", "", str(s).lower())

def load_record(idx):
    """Load record at given index into UI"""
    global current_index
    if not data or idx < 0 or idx >= len(data):
        return
    
    rec = data[idx]
    for f in FIELDS:
        entry = entries.get(f)
        if entry:
            entry.delete(0, tk.END)
            val = rec.get(f)
            entry.insert(0, "" if val is None else str(val))
    
    current_index = idx
    record_number_var.set(f"{idx + 1}/{len(data)}")

search_counter_var = tk.StringVar(value="")

def do_search():
    """Perform search in selected field"""
    global search_results, search_index
    
    key = search_field_var.get()
    val = normalize(search_value_var.get().strip())
    
    if not val:
        messagebox.showinfo("Pretraga", "Unesite vrednost za pretragu!")
        return
    
    search_results = [
        i for i, rec in enumerate(data)
        if val in normalize(rec.get(key))
    ]
    
    if search_results:
        search_index = 0
        load_record(search_results[search_index])
        search_counter_var.set(f"{search_index + 1}/{len(search_results)}")
    else:
        search_index = 0
        search_counter_var.set("")
        messagebox.showinfo("Pretraga", "Ništa nije pronađeno!")

def next_result():
    """Navigate to next search result"""
    global search_index
    if not search_results:
        return
    search_index = (search_index + 1) % len(search_results)
    load_record(search_results[search_index])
    search_counter_var.set(f"{search_index + 1}/{len(search_results)}")

# Bind Enter key to search
search_value_entry.bind('<Return>', lambda e: do_search())

ttk.Button(inner_frame, text="Traži", command=do_search).pack(side=tk.LEFT, padx=(15, 5))
ttk.Button(inner_frame, text="Sledeći", command=next_result).pack(side=tk.LEFT, padx=5)
ttk.Label(inner_frame, textvariable=search_counter_var, width=8, 
          anchor="center").pack(side=tk.LEFT, padx=5)

# --- Record Frame ---
record_frame = ttk.Frame(root, padding=(10, 10), relief="flat", borderwidth=1)
record_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=0)

entries = {}

def make_entry(parent, label_text, field_name, width=30, ipady=0):
    """Create labeled entry field"""
    frame = ttk.Frame(parent)
    frame.pack(fill=tk.X, pady=4)
    ttk.Label(frame, text=label_text, width=15, anchor="w").pack(side=tk.LEFT)
    e = ttk.Entry(frame, width=width)
    e.pack(side=tk.LEFT, fill=tk.X, ipady=ipady, ipadx=5)
    entries[field_name] = e
    return e

# --- Section 1: Basic Info ---
section1 = ttk.Frame(record_frame, padding=10, relief="groove", borderwidth=1)
section1.pack(fill=tk.X, pady=0)

id_crtez_frame = ttk.Frame(section1)
id_crtez_frame.pack(fill=tk.X, pady=5)

# ID Broj
ttk.Label(id_crtez_frame, text="ID Broj:", width=15, anchor="w").pack(side=tk.LEFT)
entries["IDENTBROJ"] = ttk.Entry(id_crtez_frame, width=20)
entries["IDENTBROJ"].pack(side=tk.LEFT, padx=(0, 18), ipady=4)

# Broj crteža
ttk.Label(id_crtez_frame, text="Broj crteža:", width=10, anchor="w").pack(side=tk.LEFT)
entries["CRTEZBROJ"] = ttk.Entry(id_crtez_frame, width=35)
entries["CRTEZBROJ"].pack(side=tk.LEFT, padx=(0, 10), ipady=4)

make_entry(section1, "Naziv dela:", "NAZIVDELA", width=68, ipady=2)
make_entry(section1, "Tehnički podaci:", "TEHNPODACI", width=68, ipady=2)

# Open button
def open_drawing():
    """Open drawing file in default image viewer"""
    crtez = entries["CRTEZBROJ"].get().strip()
    if not crtez:
        messagebox.showinfo("Otvori crtež", "Broj crteža nije specificiran!")
        return
    
    # Clean filename
    filename = crtez.replace("/", "-").replace("\\", "-") + ".jpg"
    path = os.path.join(DRAWINGS_FOLDER, filename)
    
    if os.path.exists(path):
        try:
            webbrowser.open(os.path.abspath(path))
        except Exception as e:
            messagebox.showerror("Greška", f"Nije moguće otvoriti fajl: {str(e)}")
    else:
        messagebox.showwarning("Otvori crtež", 
                              f"Fajl nije pronađen:\n{filename}\n\nProverite da li fajl postoji u folderu '{DRAWINGS_FOLDER}'")

open_btn = ttk.Button(id_crtez_frame, text="Otvori Crtež", command=open_drawing)
open_btn.pack(side=tk.LEFT, padx=10, ipadx=15)

# --- Section 2: Catalog Info ---
section2 = ttk.Frame(record_frame, padding=10, relief="groove", borderwidth=1)
section2.pack(fill=tk.X, pady=10)

make_entry(section2, "Kataloški broj:", "KATALBROJ", width=25)
make_entry(section2, "Format:", "FORMAT", width=10)
make_entry(section2, "Arhiva:", "ARHIVA", width=10)

# --- Section 3: Split - Left entries / Right buttons ---
section3 = ttk.Frame(record_frame, padding=10, relief="groove", borderwidth=1)
section3.pack(fill=tk.BOTH, expand=True, pady=10)

left_frame = ttk.Frame(section3)
left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

make_entry(left_frame, "Komentar:", "KOMENTAR", width=50)
make_entry(left_frame, "Objekat1:", "OBJEKAT1", width=50)
make_entry(left_frame, "Objekat2:", "OBJEKAT2", width=50)
make_entry(left_frame, "Objekat3:", "OBJEKAT3", width=50)
make_entry(left_frame, "Objekat4:", "OBJEKAT4", width=50)
make_entry(left_frame, "Katalog:", "KATALOG", width=50)
make_entry(left_frame, "Magsifra:", "MAGSIFRA", width=50)

right_frame = ttk.Frame(section3)
right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=15, pady=5)

def add_record():
    """Prepare UI for adding new record"""
    global current_index, search_results
    
    # Clear all entry fields
    for f in FIELDS:
        entries[f].delete(0, tk.END)

    if not data:
        current_index = 0  # First record
    else:
        current_index = len(data)  # New record at end
    
    record_number_var.set(f"{current_index + 1}/{len(data) + 1}")
    search_results = []

def save_record():
    """Save current record to data and Excel file"""
    global current_index, data

    if current_index is None:
        return
    
    # Determine if this is a new record or existing
    if current_index >= len(data):
        # New record at the end
        rec = {}
        data.append(rec)
    else:
        # Existing record
        rec = data[current_index]
    
    # Handle IDENTBROJ specially (must be integer or None)
    ident_str = entries["IDENTBROJ"].get().strip()
    if ident_str == "":
        rec["IDENTBROJ"] = None
    else:
        try:
            rec["IDENTBROJ"] = int(ident_str)
        except ValueError:
            messagebox.showerror("Greška", "ID Broj mora biti broj!")
            return

    # Save all other fields
    for f in FIELDS:
        if f == "IDENTBROJ":
            continue
        val = entries[f].get().strip()
        rec[f] = None if val == "" else val

    # Save to Excel
    success, error_msg = save_excel_data(EXCEL_FILE, data, FIELDS)
    
    if success:
        messagebox.showinfo("Sačuvaj", "Unos sačuvan!")
        # Update display
        record_number_var.set(f"{current_index + 1}/{len(data)}")
    else:
        messagebox.showerror("Greška - Fajl je zaključan", error_msg)

def delete_record():
    """Delete current record from data and Excel file"""
    global current_index, search_results, data
    
    if not data or current_index is None or current_index >= len(data):
        return
    
    crtez_broj = data[current_index].get("CRTEZBROJ", "")
    if crtez_broj:
        message = f"Da li ste sigurni da želite da izbrišete crtež {crtez_broj}?"
    else:
        message = "Da li ste sigurni da želite da izbrišete ovaj unos?"

    confirm = messagebox.askyesno("BRISANJE", message)
    if not confirm:
        return

    # Save the record in case we need to restore it
    rec = data[current_index].copy()
    del data[current_index]
    
    # Save to Excel
    success, error_msg = save_excel_data(EXCEL_FILE, data, FIELDS)
    
    if not success:
        # Restore the deleted record if save failed
        data.insert(current_index, rec)
        messagebox.showerror("Greška - Fajl je zaključan", error_msg)
        return
    
    messagebox.showinfo("Brisanje", "Unos izbrisan!")

    # Reset current index and display
    search_results = []  # Clear search results
    
    if len(data) == 0:
        current_index = None
        for e in entries.values():
            e.delete(0, tk.END)
        record_number_var.set("0/0")
    else:
        if current_index >= len(data):
            current_index = len(data) - 1
        load_record(current_index)

def make_backup():
    """Create a backup copy of the Excel file with timestamp"""
    try:
        # Create backup folder if it doesn't exist
        if not os.path.exists(BACKUP_FOLDER):
            os.makedirs(BACKUP_FOLDER)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = f"BAZACRTEZA_{timestamp}.xlsx"
        backup_path = os.path.join(BACKUP_FOLDER, backup_filename)
        
        # Copy the file
        shutil.copy2(EXCEL_FILE, backup_path)
        
        messagebox.showinfo("Rezervna kopija", 
                          f"Rezervna kopija je uspešno kreirana!\n\n{backup_filename}")
    
    except Exception as e:
        messagebox.showerror("Greška", f"Greška pri kreiranju rezervne kopije:\n{str(e)}")

# Action buttons
save_button = ttk.Button(right_frame, text="Sačuvaj Unos", command=save_record)
save_button.pack(pady=5, fill=tk.X, ipadx=10)

add_button = ttk.Button(right_frame, text="Dodaj Novi Unos", command=add_record)
add_button.pack(pady=5, fill=tk.X, ipadx=10)

delete_button = ttk.Button(right_frame, text="Izbriši Unos", command=delete_record)
delete_button.pack(pady=5, fill=tk.X, ipadx=10)

backup_button = ttk.Button(right_frame, text="Napravi rezervnu kopiju", command=make_backup)
backup_button.pack(pady=5, fill=tk.X, ipadx=10)

# --- Navigation Frame ---
nav_frame = ttk.Frame(root, padding=15)
nav_frame.pack(side="bottom", fill=tk.X, padx=10, pady=10)

def first_record():
    """Go to first record"""
    if data:
        load_record(0)

def prev_record():
    """Go to previous record"""
    global current_index
    if data and current_index is not None and current_index > 0:
        load_record(current_index - 1)

def goto_record():
    """Go to specific record number"""
    if not data:
        return
    try:
        idx = int(record_number_entry.get()) - 1
        if 0 <= idx < len(data):
            load_record(idx)
        else:
            messagebox.showwarning("Greška", 
                                  f"Unesite broj između 1 i {len(data)}")
    except ValueError:
        messagebox.showwarning("Greška", "Unesite validan broj!")

def next_record_nav():
    """Go to next record"""
    global current_index
    if data and current_index is not None and current_index < len(data) - 1:
        load_record(current_index + 1)

def last_record():
    """Go to last record"""
    if data:
        load_record(len(data) - 1)

ttk.Button(nav_frame, text="<< Prvi", command=first_record).pack(side=tk.LEFT, padx=5)
ttk.Button(nav_frame, text="< Prethodni", command=prev_record).pack(side=tk.LEFT, padx=5)
ttk.Label(nav_frame, text="Unos:").pack(side=tk.LEFT, padx=5)
record_number_entry = ttk.Entry(nav_frame, width=12, textvariable=record_number_var, 
                                justify='center')
record_number_entry.pack(side=tk.LEFT, ipady=5)
record_number_entry.bind('<Return>', lambda e: goto_record())
ttk.Button(nav_frame, text="Idi", command=goto_record).pack(side=tk.LEFT, padx=5)
ttk.Button(nav_frame, text="Sledeći >", command=next_record_nav).pack(side=tk.LEFT, padx=5)
ttk.Button(nav_frame, text="Poslednji >>", command=last_record).pack(side=tk.LEFT, padx=5)

# --- Initialize ---
if data:
    load_record(0)
else:
    record_number_var.set("0/0")

root.mainloop()